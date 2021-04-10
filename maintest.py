
from timeit import default_timer as timer
from functools import partial
import tensorflow as tf
import openpyxl
import calendar as cal

from win32com.client import Dispatch
from collections import deque
import numpy as np
import cv2
import os
import pyrebase
from google.cloud import firestore
from kivy.uix.image import Image
from kivy.properties import ObjectProperty, Clock
from kivymd.uix.screen import MDScreen
from datetime import date, datetime, timedelta
from kivymd.uix.list import  OneLineIconListItem
from kivymd.utils import asynckivy
from mtcnn import MTCNN
from kivymd.app import MDApp
from kivy.clock import Clock
from kivy.lang import Builder
from kivy.properties import StringProperty
from kivymd.toast.kivytoast import toast

firebaseConfig = {
    "apiKey": "AIzaSyBgz4SEWAvGMntcdV38nBujX6LsPeyZcRk",
    "authDomain": "instantattenddb.firebaseapp.com",
    "databaseURL": "https://instantattenddb.firebaseio.com",
    "projectId": "instantattenddb",
    "storageBucket": "instantattenddb.appspot.com",
    "messagingSenderId": "325679731114",
    "appId": "1:325679731114:web:e6846bdb4ce0eedcabdbd3",
    "measurementId": "G-JCN25C02CB"
}

from goprocam import GoProCamera, constants
import firebase_admin
from firebase_admin import credentials
from firebase_admin import firestore

# Use the application default credentials
cred = credentials.Certificate('firebase-sdk.json')
fb = firebase_admin.initialize_app(cred)
db = firestore.client()
UserUID = ""
fbu = pyrebase.initialize_app(firebaseConfig)
authintication = fbu.auth()
imgage = None
inf = []
detector = MTCNN()
tfl_file = "facenet.tflite"
def load_tflite_model(file):
    # Load the TFLite model and allocate tensors.
    interpreter = tf.lite.Interpreter(model_path=file)
    interpreter.allocate_tensors()
    return interpreter
tflite_model =load_tflite_model (tfl_file)

##exporting to excel


def export():
    if os.path.isfile("output.txt"):
        # varaibles added by user

        sectionnumber = 56125
        termstartdate = date(2020, 8, 29)
        termenddate = date(2020, 12, 11)
        # // 0 monday,...
        sectionweekdays = [6, 1, 3]
        datestack = deque()
        list_of_students = []
        today = date(2020, 11, 10)
        todays = today.strftime("%m/%d/%Y")
        print(todays)
        # today = date.today()
        with open("output.txt", "r") as x:
            attended = x.read().split(",")
        try:
            f = openpyxl.load_workbook('Att' + str(sectionnumber) + '.xlsx')
            print("file exists")

            pass
        except:
            print("exception")
            f = openpyxl.load_workbook('draft.xlsx')
            sheet = f['lecture']

            def dr(s, e):
                for n in range(int((e - s).days)):
                    yield s + timedelta(n)

            for s in dr(termstartdate, termenddate):
                if cal.weekday(s.year, s.month, s.day) in sectionweekdays:
                    datestack.append(s.strftime("%m/%d/%Y"))
                    print(s.strftime("%Y-%m-%d"))

            for c in range(4, 60, 1):
                cell = sheet.cell(7, column=c)
                if ((cell.value != None) | (len(datestack) == 0)):
                    return
                else:
                    cell.value = datestack.popleft()

            f.save('Att' + str(sectionnumber) + '.xlsx')
            f = openpyxl.load_workbook('Att' + str(sectionnumber) + '.xlsx')
            print("eceptin")
            pass

        sheet = f["Lecture"]
        today_col = -1
        for c in range(4, 60, 1):
            cell = sheet.cell(7, column=c)
            # print(cell.value )
            t = cell.value
            # print(todays)
            if (t == todays):
                todays_col = c
                # print("found")
                break

            if (cell.value == None):
                todays_col = -1
                break
                # print(cell.value.strftime("%d-%m-%Y"))

        if (sheet.cell(11, 2).value != None):
            for r in range(11, 100, 1):
                cell = "{}{}".format("B", r)
                # if null break
                if (sheet[cell].value == None):
                    break
                list_of_students.append(sheet[cell].value)

            print(attended)
            print(list_of_students)

            # print(todays)

            if (todays_col == -1):
                print("no lecture today")
            else:
                for x in attended:
                    i = 0
                    place = 0
                    for y in list_of_students:
                        if (x == y):
                            place = i + 11
                            sheet.cell(place, todays_col).value = 1
                            #print(place)
                            break
                        i = i + 1
                    #
                for r in range(11, len(list_of_students) + 11, 1):
                    cell = sheet.cell(r, todays_col)
                    # if null break
                    if (cell.value == None):
                        sheet.cell(r, todays_col).value = 0

                print("atending")
                os.remove('Att' + str(sectionnumber) + '.xlsx')
                f.save('Att' + str(sectionnumber) + '.xlsx')
        else:
            print("no students in section")


    else:
        print("attendance error")


##recognition
def pre_process(face, required_size=(160, 160)):
    ret = cv2.resize(face, required_size)
    # ret = cv2.cvtColor(ret, cv2.COLOR_BGR2RGB)
    ret = ret.astype('float32')
    # standardize pixel values across channels (global)
    mean, std = ret.mean(), ret.std()
    ret = (ret - mean) / std

    return ret


def read_image(file):
    img = cv2.imread(file)
    img = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
    return img


def load_tflite_model(file):
    # Load the TFLite model and allocate tensors.
    interpreter = tf.lite.Interpreter(model_path=file)
    interpreter.allocate_tensors()
    return interpreter


def predict(face_model, samples):
    # Get input and output tensors.
    input_details = face_model.get_input_details()
    output_details = face_model.get_output_details()

    # Test the model on random input data.
    input_shape = input_details[0]['shape']

    # input_data = np.array(np.random.random_sample(input_shape), dtype=np.float32)
    outputs = []
    for sample in samples:
        input_data = sample.reshape(input_shape)
        # input_data = np.expand_dims(input_data, axis=0)
        face_model.set_tensor(input_details[0]['index'], input_data)
        face_model.invoke()
        # The function `get_tensor()` returns a copy of the tensor data.
        # Use `tensor()` in order to get a pointer to the tensor.
        output_data = face_model.get_tensor(output_details[0]['index'])
        # print(output_data)
        outputs.append(output_data)
    ret = np.stack(outputs)
    return ret


def recognize():
    studentid = []
    refimgs = []
    studentsinsection = 0
    countsec = 0
    for count, filename in enumerate(os.listdir("sections/56125/")):
        studentid.append(filename[0:9])
        filename = "sections/56125/" + filename
        # print(studentid)
        # print(filename)
        img = read_image(filename)
        refimgs.append(img)
        studentsinsection = studentsinsection + 1
    # print(refimgs)
    print(studentsinsection)
    print(studentid)
    for count, filename in enumerate(os.listdir("ready")):
        filename = "ready/" + filename
        img = read_image(filename)
        countsec = countsec + 1
        refimgs.append(img)
    # print(attendedimgs)

    samples = [pre_process(i) for i in refimgs]

    #
    # bill1 = read_image("ready/IMG0.png")
    # x1 = read_image("sections/56125/437100230.png")
    # x2 = read_image("sections/56125/437102089.png")
    # x3 = read_image("sections/56125/437106714.png")
    # imgs = [x1, x2, x3, bill1]
    # samples = [pre_process(i) for i in imgs]


    start = timer()
    embeddings = predict(tflite_model, samples)

    f = open("output.txt", "w")
    f.write(" ")
    f.close()
    f = open("output.txt", "a")
    studentsrecognized= 0
    for i in range(studentsinsection, len(refimgs)):
        #print(i)
        for k in range(studentsinsection):
            #print(k)
            # print(np.linalg.norm(embeddings[i, :] - embeddings[k, :]))
            prediction = np.linalg.norm(embeddings[i, :] - embeddings[k, :])
            if (prediction < 0.8):
                #print(studentid[k])
                f.write(',' + studentid[k])
                studentsrecognized+=1
                break

    end = timer()
    print("FaceNet time = " + str(end - start))
    print("Total Students Recognized = " + str(studentsrecognized))
    f.close()


class loggedinUser():
    def __init__(self, UserID, uEmail):
        self.UserUID = UserID
        self.Email = uEmail
        self.db = fb.database()
        self.Sections = None
        self.UserDoc = None

    def update(self, UserID, uEmail, S,D):
        self.UserUID = UserID
        self.Email = uEmail
        self.db = fb.database()
        self.Sections = S
        self.UserDoc = D
        pass


def changeScreen(self, x):
    # login == 0
    # home == 1
    self.x = x
    if (self.x == 0):
        self.manager.current == 'Login1'
    else:
        if (self.x == 1):
            self.manager.current = 'HomeScreen1'
        else:
            if (self.x == 2):
                self.manager.current = 'CameraView1'
            else:
                if (self.x == 3):
                    self.manager.current = 'Done1'
                else:
                    self.manager.current = 'HomeScreen1'


class Login(MDScreen):
    email = ObjectProperty(None)
    password = ObjectProperty(None)

    #
    # loginbutton = ObjectProperty(None)

    def try_login(self):
        print("trying to login")
        email = self.email.text
        password = self.password.text

        if (email != None):
            if (password != None):
                try:
                    logedin = authintication.sign_in_with_email_and_password(email, password)
                    x = authintication.get_account_info(logedin['idToken'])
                    y = x['users']
                    UserUID = y[0]['localId']
                    uEmail = y[0]['email']

                    print("User: ")
                    print(x)
                    print( "UID: ")
                    print(UserUID)
                    if (UserUID != None):
                        f = open("sections.txt", "w")
                        f.write("")
                        f.close()
                        f = open("sections.txt", "a")

                        self.UserDoc = db.collection("Users").document(UserUID).get().to_dict()
                        print(self.UserDoc)
                        self.Sections =self.UserDoc["Sections"]
                        print(self.Sections)
                        for  val in self.Sections:
                            f.write(val+",")
                        f.close()
                        changeScreen(self, 1)

                except Exception as identifier:
                    # MDApp.get_running_app().self.toast_show(self,"error")
                    # self.root.app.toast_show(self,"error")
                    print(identifier)


class ItemForList(OneLineIconListItem):
    icon = StringProperty()


def takeattendance(self):
    changeScreen(self, 2)
    pass


class HomeScreen(MDScreen):
    hourc = StringProperty()
    hourc = datetime.now().strftime("%H:%M")
    datec = StringProperty()
    datec = date.today().strftime("%d/%m/%Y")
    UserUID = None
    uEmail = None

    def take_attendance(self,sectionatindex,k):
        changeScreen(self, 2)
        #print(k)
        print(sectionatindex)

    def set_list(self):
        with open("sections.txt", "r") as x:
            s = x.read().split(",")
        for i in range(0,len(s),3):
            pass
        print(s)
        async def set_list():
            for key in range(0,len(s)-1):
                print(s[key])
                await asynckivy.sleep(0)
                self.ids.box.add_widget(
                ItemForList(text=s[key], on_press = partial(self.take_attendance,key)))

        asynckivy.start(set_list())

    def refresh_callback(self, *args):
        def refresh_callback(interval):
            self.ids.box.clear_widgets()
            self.set_list()
            self.ids.refresh_layout.refresh_done()
            self.tick = 0
        Clock.schedule_once(refresh_callback, 1)

class Done(MDScreen):
    def gotohome(self):
        changeScreen(self, 1)

    def viewexcel(self):
        path = "C:/Users/qna9/Desktop/GPGUI"

        x = Dispatch("Excel.Application")
        x.Visible = True  # otherwise excel is hidden

        # newest excel does not accept forward slash in path
        wb = x.Workbooks.Open(r"C:/Users/qna9/Desktop/GPGUI/Att56125.xlsx")
        changeScreen(self, 3)
        return


class CameraView(MDScreen):
    def detectfaces(self):
        filename ="t10.png"
        img = cv2.imread(filename)
        start = timer()
        faces = detector.detect_faces(img)
        end = timer()
        print("MTCNN time = "+str(end - start))
        noface = len(faces)
        print("Number of detected faces = "+str(noface))
        i = 0
        for x in faces:
            bounding_box = x['box']
            crop_img = img[bounding_box[1]:bounding_box[1] + bounding_box[3],
                       bounding_box[0]:(bounding_box[0] + bounding_box[2])]
            face_resize = cv2.resize(crop_img, (180, 180))
            status = cv2.imwrite("ready/IMG" + str(i) + ".png", face_resize)
            #print("Image written to file-system : ", status)
            i = i + 1
        recognize()
        #export()
        changeScreen(self, 3)
        pass

    def capture(self):
        #camera = self.ids['camera']
        takax = ObjectProperty(None)
        takan = ObjectProperty(None)


        gpc = GoProCamera.GoPro(constants.auth)
        gpc.mode(constants.Hero3Commands.Mode.PhotoMode)
        gpc.delete(option="all")
        gpc.downloadLastMedia(gpc.take_photo(), custom_filename="x.jpg")
        print("we took an image")

        file = "x.jpg"
        wimg = Image(source=file, width=480, allow_stretch=True)
        self.inside.add_widget(wimg)



    def detectf(self):
        self.detectfaces()
    pass




class AttendanceApp(MDApp):

    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        # self.homescreen = HomeScreen()

    def build(self):
        self.root = Builder.load_file("AttendanceApp.kv")
        self.theme_cls.primary_palette = "Cyan"
        return self.root

    def toast_print(self, text):

        toast(text)


if __name__ == "__main__":
    AttendanceApp().run()
